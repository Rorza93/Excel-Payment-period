from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict

# Load the workbook and select the active worksheet
wb = load_workbook('test.xlsm')
ws = wb.active

# Dictionary to hold payments, grouped by account numbers
payments_by_account = defaultdict(list)

for row in ws.iter_rows(min_row=2, values_only=True):
    account_number, payment_date = row[0], row[1]

    # Skip rows where the account number or date is missing
    if not account_number or not payment_date:
        continue

    # Add the payment date to the respective account's list
    payments_by_account[account_number].append(payment_date)

# Dictionary to hold the date differences and frequency for each account
account_info = {}

# Iterate through each account and calculate the date difference
for account, dates in payments_by_account.items():
    # Sort the dates to find the most recent date
    sorted_dates = sorted(dates)
    last_payment_date = sorted_dates[-1]

    # If there's more than one date, calculate the frequency
    if len(sorted_dates) > 1:
        second_last_payment_date = sorted_dates[-2]
        date_diff = (last_payment_date - second_last_payment_date).days

        if date_diff >= 15 and date_diff < 45:
            frequency = '1 Month'
        elif date_diff >= 45 and date_diff < 75:
            frequency = '2 Months'
        elif date_diff >= 75 and date_diff < 105:
            frequency = '3 Months'
        elif date_diff >= 105 and date_diff < 135:
            frequency = '4 Months'
        else:
            frequency = 'Other or irregular period'

        # Calculate the next expected payment date based on the frequency and add 20 days
        if 'Month' in frequency:
            num_months = int(frequency.split()[0])
            expected_next_payment = last_payment_date + timedelta(days=num_months * 30 + 20)
        else:
            expected_next_payment = 'Irregular'

        # Store the frequency and next expected payment date
        account_info[account] = {
            'frequency': frequency,
            'next_payment': expected_next_payment,
            'last_payment': last_payment_date
        }
    else:
        account_info[account] = {
            'frequency': 'Only one payment',
            'next_payment': 'N/A',
            'last_payment': last_payment_date
        }

# Optional: Print the info to verify
for account, info in account_info.items():
    #print(f"Account {account} - Payment period: {info['frequency']}, Last Payment: {info['last_payment'].strftime('%Y-%m-%d')}, Next expected payment: {info['next_payment'] if isinstance(info['next_payment'], datetime) else info['next_payment']}")

    # Check if the expected next payment date has passed
    if isinstance(info['next_payment'], datetime) and info['next_payment'] < datetime.now():
        print(f"Account {account} - Payment period: {info['frequency']}, Last Payment: {info['last_payment'].strftime('%Y-%m-%d')}, Next expected payment: {info['next_payment'] if isinstance(info['next_payment'], datetime) else info['next_payment']}")
        print(f"Account {account} - Next expected payment is overdue!")
    #elif isinstance(info['next_payment'], datetime):
     #   print(f"Account {account} - Next expected payment is upcoming.")
